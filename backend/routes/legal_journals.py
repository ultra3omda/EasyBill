from fastapi import APIRouter, HTTPException, status, Depends, Query
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
from datetime import datetime, timezone
from io import BytesIO
import os
from typing import Optional, List
from utils.dependencies import get_current_user, get_current_company

router = APIRouter(prefix="/api/legal-journals", tags=["Legal Journals"])

mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

JOURNAL_TYPES = {
    "sales": {
        "label": "Journal des Ventes",
        "description": "Factures de vente, avoirs clients",
        "account_prefixes": ["70", "41", "43"],
        "document_types": ["invoice", "credit_note"],
        "journal_types": ["sales", "ventes"],
        "icon": "TrendingUp",
        "color": "green"
    },
    "purchases": {
        "label": "Journal des Achats",
        "description": "Factures fournisseurs, avoirs fournisseurs",
        "account_prefixes": ["60", "40", "43"],
        "document_types": ["supplier_invoice"],
        "journal_types": ["purchases", "achats"],
        "icon": "ShoppingCart",
        "color": "blue"
    },
    "cash": {
        "label": "Journal de Caisse",
        "description": "Opérations de caisse (espèces)",
        "account_prefixes": ["53"],
        "document_types": ["cash_payment", "cash_expense"],
        "journal_types": ["cash", "caisse"],
        "icon": "Banknote",
        "color": "yellow"
    },
    "bank": {
        "label": "Journal de Banque",
        "description": "Virements, chèques, opérations bancaires",
        "account_prefixes": ["52"],
        "document_types": ["bank_transfer", "check"],
        "journal_types": ["bank", "banque"],
        "icon": "Building2",
        "color": "indigo"
    },
    "od": {
        "label": "Journal des OD",
        "description": "Opérations diverses, écritures d'inventaire, régularisations",
        "account_prefixes": ["1", "2", "6", "7"],
        "document_types": [],
        "journal_types": ["general", "od", "inventory"],
        "icon": "RefreshCw",
        "color": "purple"
    }
}


def serialize_journal_entry(e: dict) -> dict:
    lines = []
    for line in e.get("lines", []):
        lines.append({
            "account_code": line.get("account_code"),
            "account_name": line.get("account_name"),
            "debit": round(line.get("debit", 0), 3),
            "credit": round(line.get("credit", 0), 3),
            "description": line.get("description", "")
        })

    computed_debit = round(sum(l.get("debit", 0) for l in e.get("lines", [])), 3)
    computed_credit = round(sum(l.get("credit", 0) for l in e.get("lines", [])), 3)
    total_debit = e.get("total_debit") or computed_debit
    total_credit = e.get("total_credit") or computed_credit
    entry_number = e.get("entry_number") or e.get("reference") or ""

    return {
        "id": str(e["_id"]),
        "entry_number": entry_number,
        "date": e.get("date").isoformat() if e.get("date") else None,
        "reference": e.get("reference"),
        "description": e.get("description"),
        "journal_type": e.get("journal_type", "general"),
        "document_type": e.get("document_type"),
        "document_id": str(e.get("document_id")) if e.get("document_id") else None,
        "lines": lines,
        "total_debit": total_debit,
        "total_credit": total_credit,
        "status": e.get("status", "draft"),
        "created_at": e.get("created_at").isoformat() if e.get("created_at") else None,
    }


def _build_journal_query(company_id: str, journal_key: str,
                          date_from: Optional[str], date_to: Optional[str]) -> dict:
    """Build MongoDB query for a given legal journal type."""
    conf = JOURNAL_TYPES.get(journal_key)
    if not conf:
        raise HTTPException(status_code=404, detail="Journal type not found")

    query: dict = {"company_id": ObjectId(company_id)}

    # Date filter
    date_filter = {}
    if date_from:
        date_filter["$gte"] = datetime.fromisoformat(date_from)
    if date_to:
        date_filter["$lte"] = datetime.fromisoformat(date_to + "T23:59:59")
    if date_filter:
        query["date"] = date_filter

    # Journal type OR document type filter
    or_conditions = []
    if conf["journal_types"]:
        or_conditions.append({"journal_type": {"$in": conf["journal_types"]}})
    if conf["document_types"]:
        or_conditions.append({"document_type": {"$in": conf["document_types"]}})

    # For OD: exclude known other types
    if journal_key == "od":
        query["journal_type"] = {"$in": ["general", "od", "inventory"]}
        query["document_type"] = {"$nin": ["invoice", "credit_note", "supplier_invoice", "cash_payment", "cash_expense", "bank_transfer", "check"]}
    elif or_conditions:
        query["$or"] = or_conditions

    return query


@router.get("/types")
async def list_journal_types():
    """Return all available legal journal types."""
    return [
        {"key": k, **{f: v for f, v in v.items() if f != "account_prefixes" and f != "journal_types" and f != "document_types"}}
        for k, v in JOURNAL_TYPES.items()
    ]


@router.get("/summary")
async def get_journals_summary(
    company_id: str = Query(...),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get a summary of all legal journals (entry count + totals per journal)."""
    await get_current_company(current_user, company_id)

    result = []
    for key, conf in JOURNAL_TYPES.items():
        try:
            query = _build_journal_query(company_id, key, date_from, date_to)
        except Exception:
            continue

        pipeline = [
            {"$match": query},
            {"$group": {
                "_id": None,
                "count": {"$sum": 1},
                "total_debit": {"$sum": "$total_debit"},
                "total_credit": {"$sum": "$total_credit"}
            }}
        ]
        agg = await db.journal_entries.aggregate(pipeline).to_list(1)
        stats = agg[0] if agg else {"count": 0, "total_debit": 0, "total_credit": 0}

        result.append({
            "key": key,
            "label": conf["label"],
            "description": conf["description"],
            "icon": conf["icon"],
            "color": conf["color"],
            "count": stats.get("count", 0),
            "total_debit": round(stats.get("total_debit", 0) or 0, 3),
            "total_credit": round(stats.get("total_credit", 0) or 0, 3),
        })

    return result


@router.get("/{journal_key}")
async def get_legal_journal(
    journal_key: str,
    company_id: str = Query(...),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    current_user: dict = Depends(get_current_user)
):
    """Get entries for a specific legal journal."""
    await get_current_company(current_user, company_id)

    query = _build_journal_query(company_id, journal_key, date_from, date_to)

    total = await db.journal_entries.count_documents(query)
    skip = (page - 1) * page_size

    entries = await db.journal_entries.find(query).sort("date", 1).skip(skip).limit(page_size).to_list(page_size)

    # Compute running balance per account
    totals_debit = 0.0
    totals_credit = 0.0
    serialized = []
    for e in entries:
        s = serialize_journal_entry(e)
        totals_debit += s["total_debit"]
        totals_credit += s["total_credit"]
        serialized.append(s)

    return {
        "journal_key": journal_key,
        "label": JOURNAL_TYPES[journal_key]["label"],
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": (total + page_size - 1) // page_size,
        "totals_debit": round(totals_debit, 3),
        "totals_credit": round(totals_credit, 3),
        "entries": serialized
    }


@router.get("/{journal_key}/export/excel")
async def export_legal_journal_excel(
    journal_key: str,
    company_id: str = Query(...),
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export a legal journal to Excel."""
    await get_current_company(current_user, company_id)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non disponible")

    conf = JOURNAL_TYPES.get(journal_key)
    if not conf:
        raise HTTPException(status_code=404, detail="Journal type not found")

    query = _build_journal_query(company_id, journal_key, date_from, date_to)
    entries = await db.journal_entries.find(query).sort("date", 1).to_list(5000)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = conf["label"][:31]

    header_fill = PatternFill("solid", fgColor="6366F1")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill("solid", fgColor="E0E7FF")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = conf["label"]
    ws["A1"].font = Font(bold=True, size=14, color="3730A3")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    period = ""
    if date_from and date_to:
        period = f"Période : {date_from} au {date_to}"
    elif date_from:
        period = f"Du {date_from}"
    elif date_to:
        period = f"Au {date_to}"
    ws["A2"] = period
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["N° Écriture", "Date", "Référence", "Description / Compte", "Code", "Intitulé", "Débit", "Crédit"]
    ws.append([])
    row_idx = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin

    # Data rows
    total_d = 0.0
    total_c = 0.0
    row_idx = 5

    for e in entries:
        s = serialize_journal_entry(e)
        date_str = s["date"][:10] if s.get("date") else ""
        lines = s.get("lines", [])

        if not lines:
            ws.append([s["entry_number"], date_str, s["reference"], s["description"], "", "", s["total_debit"], s["total_credit"]])
            row_idx += 1
            continue

        for i, line in enumerate(lines):
            entry_num = s["entry_number"] if i == 0 else ""
            date_v = date_str if i == 0 else ""
            ref = s["reference"] if i == 0 else ""
            desc = s["description"] if i == 0 else ""
            row = [entry_num, date_v, ref, desc,
                   line["account_code"], line["account_name"],
                   line["debit"] if line["debit"] else None,
                   line["credit"] if line["credit"] else None]
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin
                if col in [7, 8] and val:
                    cell.number_format = '#,##0.000'
            if i == 0:
                total_d += s["total_debit"]
                total_c += s["total_credit"]
            row_idx += 1

    # Totals row
    ws.cell(row=row_idx, column=1, value="TOTAL").font = Font(bold=True)
    cell_d = ws.cell(row=row_idx, column=7, value=round(total_d, 3))
    cell_d.font = Font(bold=True)
    cell_d.number_format = '#,##0.000'
    cell_c = ws.cell(row=row_idx, column=8, value=round(total_c, 3))
    cell_c.font = Font(bold=True)
    cell_c.number_format = '#,##0.000'
    for col in range(1, 9):
        ws.cell(row=row_idx, column=col).fill = subheader_fill

    # Column widths
    col_widths = [14, 12, 14, 35, 10, 30, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"journal_{journal_key}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
